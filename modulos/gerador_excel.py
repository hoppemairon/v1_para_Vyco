import pandas as pd
import os
from datetime import datetime
from typing import Dict, List
import streamlit as st
import io

class GeradorExcel:
    """Classe responsável por gerar os arquivos Excel no formato Vyco"""
    
    def __init__(self):
        self.diretorio_modelos = "modelos_vyco"
        self.diretorio_preenchimento = "preenchimento_vyco"
        self.arquivos_gerados = []
    
    def gerar_todas_planilhas(self, dados_processados: Dict[str, pd.DataFrame], 
                             diretorio_saida: str) -> List[str]:
        """
        Gera todas as planilhas no formato Vyco
        
        Args:
            dados_processados: Dicionário com todos os dados processados
            diretorio_saida: Diretório onde salvar as planilhas
            
        Returns:
            List[str]: Lista com nomes dos arquivos gerados
        """
        try:
            # Criar diretório de saída se não existir
            os.makedirs(diretorio_saida, exist_ok=True)
            
            self.arquivos_gerados = []
            
            # Mapear dados para arquivos
            mapeamento_arquivos = {
                'categorias': 'Cadastros - Categorias.xlsx',
                'centros_custo': 'Cadastros - Centros de Custo.xlsx',
                'contas_correntes': 'Cadastros - Contas Corrente.xlsx',
                'contatos': 'Cadastros - Contatos.xlsx',
                'lancamentos': 'Financeiro - Lançamentos.xlsx',
                'transferencias': 'Financeiro - Transferências.xlsx'
            }
            
            # Gerar cada planilha
            for tipo_dado, nome_arquivo in mapeamento_arquivos.items():
                if tipo_dado in dados_processados:
                    caminho_arquivo = os.path.join(diretorio_saida, nome_arquivo)
                    
                    # Gerar planilha específica
                    sucesso = self._gerar_planilha_especifica(
                        tipo_dado, 
                        dados_processados[tipo_dado], 
                        caminho_arquivo
                    )
                    
                    if sucesso:
                        self.arquivos_gerados.append(nome_arquivo)
                        st.success(f"✅ {nome_arquivo} gerado com sucesso!")
            
            return self.arquivos_gerados
            
        except Exception as e:
            st.error(f"Erro ao gerar planilhas: {str(e)}")
            return []
    
    def gerar_planilhas_memoria(self, dados_processados: Dict[str, pd.DataFrame]) -> Dict[str, bytes]:
        """
        Gera todas as planilhas em memória (sem salvar no disco)
        
        Args:
            dados_processados: Dicionário com todos os dados processados
            
        Returns:
            Dict[str, bytes]: Dicionário com nome do arquivo e conteúdo em bytes
        """
        try:
            arquivos_memoria = {}
            
            # Mapear dados para arquivos
            mapeamento_arquivos = {
                'categorias': 'Cadastros - Categorias.xlsx',
                'centros_custo': 'Cadastros - Centros de Custo.xlsx',
                'contas_correntes': 'Cadastros - Contas Corrente.xlsx',
                'contatos': 'Cadastros - Contatos.xlsx',
                'lancamentos': 'Financeiro - Lançamentos.xlsx',
                'transferencias': 'Financeiro - Transferências.xlsx'
            }
            
            # Gerar cada planilha em memória
            for tipo_dado, nome_arquivo in mapeamento_arquivos.items():
                if tipo_dado in dados_processados:
                    conteudo_bytes = self._gerar_planilha_memoria(
                        tipo_dado, 
                        dados_processados[tipo_dado]
                    )
                    
                    if conteudo_bytes:
                        arquivos_memoria[nome_arquivo] = conteudo_bytes
            
            return arquivos_memoria
            
        except Exception as e:
            st.error(f"Erro ao gerar planilhas em memória: {str(e)}")
            return {}
    
    def _gerar_planilha_memoria(self, tipo_planilha: str, dados: pd.DataFrame) -> bytes:
        """
        Gera uma planilha específica em memória
        
        Args:
            tipo_planilha: Tipo da planilha (categorias, lancamentos, etc.)
            dados: DataFrame com os dados
            
        Returns:
            bytes: Conteúdo da planilha em bytes
        """
        try:
            # Aplicar formatação específica para cada tipo
            dados_formatados = self._aplicar_formatacao_especifica(tipo_planilha, dados)
            
            # Ler modelo se existir para garantir estrutura correta
            dados_finais = self._aplicar_estrutura_modelo(tipo_planilha, dados_formatados)
            
            # Criar buffer em memória
            buffer = io.BytesIO()
            
            # Salvar no buffer
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                dados_finais.to_excel(writer, sheet_name='Dados', index=False)
                
                # Aplicar formatação adicional se necessário
                self._aplicar_formatacao_excel(writer, tipo_planilha)
            
            # Obter bytes do buffer
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            st.error(f"Erro ao gerar {tipo_planilha} em memória: {str(e)}")
            return b''
    
    def _gerar_planilha_especifica(self, tipo_planilha: str, 
                                  dados: pd.DataFrame, 
                                  caminho_arquivo: str) -> bool:
        """
        Gera uma planilha específica aplicando formatação adequada
        
        Args:
            tipo_planilha: Tipo da planilha (categorias, lancamentos, etc.)
            dados: DataFrame com os dados
            caminho_arquivo: Caminho onde salvar o arquivo
            
        Returns:
            bool: True se gerado com sucesso
        """
        try:
            # Aplicar formatação específica para cada tipo
            dados_formatados = self._aplicar_formatacao_especifica(tipo_planilha, dados)
            
            # Ler modelo se existir para garantir estrutura correta
            dados_finais = self._aplicar_estrutura_modelo(tipo_planilha, dados_formatados)
            
            # Salvar arquivo com tratamento de erro
            try:
                with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
                    dados_finais.to_excel(writer, sheet_name='Dados', index=False)
                    
                    # Aplicar formatação adicional se necessário
                    self._aplicar_formatacao_excel(writer, tipo_planilha)
                
                return True
            except PermissionError:
                st.error(f"❌ Erro de permissão ao salvar {caminho_arquivo}. Feche o arquivo se estiver aberto.")
                return False
            except Exception as save_error:
                st.error(f"❌ Erro ao salvar {caminho_arquivo}: {str(save_error)}")
                return False
            
        except Exception as e:
            st.error(f"Erro ao gerar {tipo_planilha}: {str(e)}")
            return False
    
    def _aplicar_formatacao_especifica(self, tipo_planilha: str, dados: pd.DataFrame) -> pd.DataFrame:
        """
        Aplica formatação específica para cada tipo de planilha
        
        Args:
            tipo_planilha: Tipo da planilha
            dados: DataFrame original
            
        Returns:
            pd.DataFrame: DataFrame formatado
        """
        dados_formatados = dados.copy()
        
        try:
            if tipo_planilha == 'categorias':
                dados_formatados = self._formatar_categorias(dados_formatados)
            elif tipo_planilha == 'centros_custo':
                dados_formatados = self._formatar_centros_custo(dados_formatados)
            elif tipo_planilha == 'contas_correntes':
                dados_formatados = self._formatar_contas_correntes(dados_formatados)
            elif tipo_planilha == 'contatos':
                dados_formatados = self._formatar_contatos(dados_formatados)
            elif tipo_planilha == 'lancamentos':
                dados_formatados = self._formatar_lancamentos(dados_formatados)
            elif tipo_planilha == 'transferencias':
                dados_formatados = self._formatar_transferencias(dados_formatados)
            
        except Exception as e:
            st.warning(f"Aviso na formatação de {tipo_planilha}: {str(e)}")
        
        return dados_formatados
    
    def _formatar_categorias(self, dados: pd.DataFrame) -> pd.DataFrame:
        """
        Formata dados de categorias com estrutura completa do Vyco
        Inclui todas as colunas necessárias e calcula códigos pai
        """
        dados_formatados = dados.copy()
        
        # Garantir que temos as colunas básicas
        if 'Codigo' not in dados_formatados.columns:
            raise ValueError("Coluna 'Codigo' não encontrada nos dados processados")
        
        if 'Nome' not in dados_formatados.columns:
            raise ValueError("Coluna 'Nome' não encontrada nos dados processados")
        
        # Renomear para usar acento conforme modelo Vyco
        dados_formatados = dados_formatados.rename(columns={'Codigo': 'Código'})
        
        # Calcular código pai baseado na hierarquia
        dados_formatados['Código Pai'] = dados_formatados['Código'].apply(self._calcular_codigo_pai)
        
        # Adicionar colunas obrigatórias do Vyco com valores padrão
        dados_formatados['Competência por Parcela'] = ''  # Vazio por padrão
        dados_formatados['Desconto %'] = 0.0  # Zero por padrão
        dados_formatados['Desconto R$'] = 0.0  # Zero por padrão
        dados_formatados['Dias para vencimento'] = 0  # Zero por padrão
        dados_formatados['Gera Boleto'] = 'Não'  # Não por padrão
        dados_formatados['Gera Nota Fiscal'] = 'Não'  # Não por padrão
        dados_formatados['Gera Recibo'] = 'Não'  # Não por padrão
        
        # Definir ordem das colunas conforme modelo Vyco
        colunas_ordenadas = [
            'Código', 'Nome', 'Código Pai', 'Competência por Parcela',
            'Desconto %', 'Desconto R$', 'Dias para vencimento',
            'Gera Boleto', 'Gera Nota Fiscal', 'Gera Recibo'
        ]
        
        # Reordenar colunas
        dados_formatados = dados_formatados[colunas_ordenadas]
        
        return dados_formatados
    
    def _calcular_codigo_pai(self, codigo: str) -> str:
        """
        Calcula o código pai baseado na hierarquia
        Exemplo: '2.1.1.1' -> '2.1.1', '2.1.1' -> '2.1', '2.1' -> '2'
        
        Args:
            codigo: Código da categoria
            
        Returns:
            str: Código do pai ou vazio se for nível raiz
        """
        if not isinstance(codigo, str):
            codigo = str(codigo)
        
        codigo = codigo.strip()
        
        # Se não contém ponto, é nível raiz (não tem pai)
        if '.' not in codigo:
            return ''
        
        # Dividir por pontos e remover o último nível
        partes = codigo.split('.')
        if len(partes) > 1:
            return '.'.join(partes[:-1])
        
        return ''
    
    def _formatar_centros_custo(self, dados: pd.DataFrame) -> pd.DataFrame:
        """Formata dados de centros de custo"""
        colunas_esperadas = ['Codigo', 'Nome', 'Ativo']
        return self._garantir_colunas(dados, colunas_esperadas)
    
    def _formatar_contas_correntes(self, dados: pd.DataFrame) -> pd.DataFrame:
        """Formata dados de contas correntes conforme modelo Vyco"""
        dados_formatados = dados.copy()
        
        # Mapear para estrutura do modelo Vyco
        if 'Codigo' in dados_formatados.columns:
            dados_formatados = dados_formatados.drop('Codigo', axis=1)
        
        # Usar DataInicial calculada ou deixar vazio
        if 'DataInicial' in dados_formatados.columns:
            dados_formatados['Data Inicial'] = dados_formatados['DataInicial']
            dados_formatados = dados_formatados.drop('DataInicial', axis=1)
        else:
            dados_formatados['Data Inicial'] = ''  # Vazio por padrão
        
        # Preservar Data Inicial original (não alterar dados do usuário)
        if 'Data Inicial' in dados_formatados.columns:
            # Apenas padronizar formato, preservar vazias
            dados_formatados['Data Inicial'] = dados_formatados['Data Inicial'].astype(str)
            dados_formatados.loc[dados_formatados['Data Inicial'].isin(['', 'nan', 'None', 'NaT']), 'Data Inicial'] = ''
        
        # Adicionar Valor Inicial
        dados_formatados['Valor Inicial'] = 0.0  # Zero por padrão
        
        # Remover coluna 'Ativo' se existir (não está no modelo)
        if 'Ativo' in dados_formatados.columns:
            dados_formatados = dados_formatados.drop('Ativo', axis=1)
        
        # Definir ordem das colunas conforme modelo
        colunas_ordenadas = ['Nome', 'Tipo', 'Data Inicial', 'Valor Inicial']
        dados_formatados = dados_formatados[colunas_ordenadas]
        
        return dados_formatados
    
    def _formatar_contatos(self, dados: pd.DataFrame) -> pd.DataFrame:
        """Formata dados de contatos conforme especificação Vyco"""
        dados_formatados = dados.copy()
        
        # Garantir que todas as colunas obrigatórias existem conforme documentação
        colunas_esperadas = [
            'Nome', 'Tipo', 'Documento', 'E-mail', 'Enviar e-mail?',
            'Telefone Residencial', 'Telefone Comercial', 'Telefone Celular',
            'Contribuinte ICMS?', 'Inscrição Estadual', 'Inscrição Municipal'
        ]
        
        # Adicionar colunas faltantes se não existirem
        for coluna in colunas_esperadas:
            if coluna not in dados_formatados.columns:
                if coluna == 'Tipo':
                    dados_formatados[coluna] = 0  # Não Identificado
                elif coluna in ['Enviar e-mail?', 'Contribuinte ICMS?']:
                    dados_formatados[coluna] = False  # Não
                else:
                    dados_formatados[coluna] = ''  # Vazio
        
        # Reordenar colunas conforme especificação
        dados_formatados = dados_formatados[colunas_esperadas]
        
        return dados_formatados
    
    def _formatar_lancamentos(self, dados: pd.DataFrame) -> pd.DataFrame:
        """Formata dados de lançamentos conforme modelo Vyco"""
        dados_formatados = dados.copy()
        
        # Garantir formato de data (PRESERVANDO dados originais)
        if 'Data' in dados_formatados.columns:
            # Tentar padronizar formato apenas para datas válidas, preservar vazias
            dados_formatados['Data'] = dados_formatados['Data'].astype(str)
            # Se estiver vazio/NaN, manter vazio
            dados_formatados.loc[dados_formatados['Data'].isin(['', 'nan', 'None', 'NaT']), 'Data'] = ''
        
        # Garantir formato de valor
        if 'Valor' in dados_formatados.columns:
            dados_formatados['Valor'] = pd.to_numeric(dados_formatados['Valor'], errors='coerce')
            dados_formatados['Valor'] = dados_formatados['Valor'].round(2)
        
        # Garantir formato de valor emissão se existir
        if 'Valor Emissão' in dados_formatados.columns:
            dados_formatados['Valor Emissão'] = pd.to_numeric(dados_formatados['Valor Emissão'], errors='coerce')
            dados_formatados['Valor Emissão'] = dados_formatados['Valor Emissão'].round(2)
        
        # Garantir formato de repetição
        if 'Repetição' in dados_formatados.columns:
            dados_formatados['Repetição'] = pd.to_numeric(dados_formatados['Repetição'], errors='coerce').fillna(0).astype(int)
        
        # Garantir formato de total parcelas
        if 'Total Parcelas' in dados_formatados.columns:
            dados_formatados['Total Parcelas'] = pd.to_numeric(dados_formatados['Total Parcelas'], errors='coerce').fillna(1).astype(int)
        
        # Colunas conforme especificação Vyco
        colunas_esperadas = [
            'Confirmado', 'Data Emissão', 'Data', 'Valor Emissão', 'Valor',
            'Repetição', 'Total Parcelas', 'Descrição', 'Categoria',
            'Centro de Custo', 'Conta Corrente', 'Contato'
        ]
        
        return self._garantir_colunas(dados_formatados, colunas_esperadas)
    
    def _formatar_transferencias(self, dados: pd.DataFrame) -> pd.DataFrame:
        """Formata dados de transferências conforme modelo Vyco"""
        dados_formatados = dados.copy()
        
        # Garantir formato de data (PRESERVANDO dados originais)
        if 'Data' in dados_formatados.columns:
            # Tentar padronizar formato apenas para datas válidas, preservar vazias
            dados_formatados['Data'] = dados_formatados['Data'].astype(str)
            # Se estiver vazio/NaN, manter vazio
            dados_formatados.loc[dados_formatados['Data'].isin(['', 'nan', 'None', 'NaT']), 'Data'] = ''
        
        # Garantir formato de valor
        if 'Valor' in dados_formatados.columns:
            dados_formatados['Valor'] = pd.to_numeric(dados_formatados['Valor'], errors='coerce')
            dados_formatados['Valor'] = dados_formatados['Valor'].round(2)
        
        # Garantir que temos todas as colunas conforme modelo Vyco
        colunas_esperadas = ['Data', 'Valor', 'Descrição', 'Conta Débito', 'Conta Crédito']
        
        # Adicionar colunas faltantes
        for coluna in colunas_esperadas:
            if coluna not in dados_formatados.columns:
                dados_formatados[coluna] = ''
        
        # Reordenar conforme modelo
        dados_formatados = dados_formatados[colunas_esperadas]
        
        return dados_formatados
    
    def _garantir_colunas(self, dados: pd.DataFrame, colunas_esperadas: List[str]) -> pd.DataFrame:
        """
        Garante que o DataFrame tenha todas as colunas esperadas
        
        Args:
            dados: DataFrame original
            colunas_esperadas: Lista de colunas que devem existir
            
        Returns:
            pd.DataFrame: DataFrame com todas as colunas
        """
        dados_final = dados.copy()
        
        # Adicionar colunas faltantes
        for coluna in colunas_esperadas:
            if coluna not in dados_final.columns:
                dados_final[coluna] = ""
        
        # Reordenar colunas
        colunas_existentes = [col for col in colunas_esperadas if col in dados_final.columns]
        dados_final = dados_final[colunas_existentes]
        
        return dados_final
    
    def _aplicar_estrutura_modelo(self, tipo_planilha: str, dados: pd.DataFrame) -> pd.DataFrame:
        """
        Aplica estrutura do modelo Vyco se existir
        
        Args:
            tipo_planilha: Tipo da planilha
            dados: DataFrame com dados
            
        Returns:
            pd.DataFrame: DataFrame com estrutura do modelo
        """
        # Mapear tipo para arquivo modelo
        mapeamento_modelos = {
            'categorias': 'Cadastros - Categorias.xlsx',
            'centros_custo': 'Cadastros - Centros de Custo.xlsx',
            'contas_correntes': 'Cadastros - Contas Corrente.xlsx',
            'contatos': 'Cadastros - Contatos.xlsx',
            'lancamentos': 'Financeiro - Lançamentos.xlsx',
            'transferencias': 'Financeiro - Transferências.xlsx'
        }
        
        if tipo_planilha not in mapeamento_modelos:
            return dados
        
        nome_arquivo_modelo = mapeamento_modelos[tipo_planilha]
        caminho_modelo = os.path.join(self.diretorio_modelos, nome_arquivo_modelo)
        
        try:
            if os.path.exists(caminho_modelo):
                # Ler estrutura do modelo
                modelo = pd.read_excel(caminho_modelo)
                
                # Usar colunas do modelo como referência
                colunas_modelo = modelo.columns.tolist()
                
                # Garantir que o DataFrame tenha as colunas do modelo
                dados_final = self._garantir_colunas(dados, colunas_modelo)
                
                return dados_final
            
        except Exception as e:
            st.warning(f"Não foi possível aplicar modelo para {tipo_planilha}: {str(e)}")
        
        return dados
    
    def _aplicar_formatacao_excel(self, writer, tipo_planilha: str) -> None:
        """
        Aplica formatação adicional ao arquivo Excel
        
        Args:
            writer: ExcelWriter object
            tipo_planilha: Tipo da planilha
        """
        try:
            worksheet = writer.sheets['Dados']
            
            # Formatação básica
            # Ajustar largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar formatação específica por tipo
            if tipo_planilha in ['lancamentos', 'transferencias']:
                # Formatação para valores monetários
                from openpyxl.styles import NamedStyle
                
                currency_style = NamedStyle(name='currency')
                currency_style.number_format = 'R$ #,##0.00'
                
                # Aplicar formatação em colunas de valor
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        if 'valor' in str(worksheet.cell(1, cell.column).value).lower():
                            cell.style = currency_style
                            
        except Exception as e:
            st.warning(f"Aviso na formatação Excel: {str(e)}")
    
    def ler_instrucoes_preenchimento(self, tipo_planilha: str) -> str:
        """
        Lê instruções de preenchimento para um tipo de planilha
        
        Args:
            tipo_planilha: Tipo da planilha
            
        Returns:
            str: Instruções de preenchimento
        """
        mapeamento_instrucoes = {
            'categorias': 'Cadastros - Categorias.txt',
            'centros_custo': 'Cadastros - Centros de Custo.txt',
            'contas_correntes': 'Cadastros - Contas Corrente.txt',
            'contatos': 'Cadastros - Contatos.txt',
            'lancamentos': 'Financeiro - Lançamentos.txt',
            'transferencias': 'Financeiro - Transferências.txt'
        }
        
        if tipo_planilha not in mapeamento_instrucoes:
            return "Instruções não disponíveis"
        
        nome_arquivo = mapeamento_instrucoes[tipo_planilha]
        caminho_arquivo = os.path.join(self.diretorio_preenchimento, nome_arquivo)
        
        try:
            if os.path.exists(caminho_arquivo):
                with open(caminho_arquivo, 'r', encoding='utf-8') as arquivo:
                    return arquivo.read()
            else:
                return f"Arquivo de instruções não encontrado: {caminho_arquivo}"
                
        except Exception as e:
            return f"Erro ao ler instruções: {str(e)}"
    
    def obter_estatisticas_geracao(self) -> Dict[str, int]:
        """
        Retorna estatísticas da geração de arquivos
        
        Returns:
            Dict: Estatísticas da geração
        """
        return {
            'total_arquivos_gerados': len(self.arquivos_gerados),
            'arquivos_gerados': self.arquivos_gerados
        }